import sys
import openai
from openpyxl import load_workbook
import pandas as pd
import re
from openai import OpenAI

if __name__ == '__main__':
    
    # Point to the local server
    #client = OpenAI(base_url="http://localhost:1234/v1", api_key="lm-studio")
    client = OpenAI(base_url="http://localhost:11434/v1", api_key="ollama")
    #client = OpenAI(base_url="http://localhost:10000/v1", api_key="vllm")
    
    # be sure that if you use lmstudio, after loading a model, you go into the 'My Models' tab (left nav)
    # and then for the model, select the way the model should 'talk'. E.g. fof llama3-7b-instruct, pick 'llama3'
    # The general prompts here are for llama3....
    
    
    MODEL = "lmstudio-community/Meta-Llama-3-8B-Instruct-GGUF" # lmstudio
    MODEL = "llama3" # ollama
    MODEL = "qwen:14b"
    #MODEL = "ruslanmv/Medical-Llama3-8B-GGUF"
    #MODEL = "TheBloke/Mixtral-8x7B-Instruct-v0.1-GGUF"  # lmstudio
    #MODEL = "meta-llama/Meta-Llama-3-70B-Instruct" #vllm on A100
    name = MODEL.split("/")[-1]
    print (name)
    
    #---
    
    
    def get_question(exam_class):
        if exam_class.lower() == "cervical spine fracture":
            # apparently by using 'is' it consideres only acute fractures. Could add 'or was' to include old fractures
            return ("Is there likely or definitely an acute fracture (displaced on non-displaced) of any part of the cervical spine including C1, C2, C3, C4, C5, C6, C7, or of the odontoid? You shoudl consider any part of the spine (the body, lateral mass, lamina, posterior elements, transverse process, spinous process, or osteophytes as part of the spine) Answer using these options: ['Yes', 'No']. ") 
        if exam_class.lower() == "pulmonary embolism":
            return ("Is there likely or definitely a pulmonary embolism, which may appear as a filling defect in a pulmnary artery, present? Options are: ['Yes', 'No']. If not specifically mentioned, then answer 'No'") 
        if exam_class.lower() == "pneumonia":
            return ("Is there concern for pneumonia or a developing opacity in the lung? Options are: ['Yes', 'No']. If not specifically mentioned, then answer 'No'") 
        if exam_class.lower() == "liver metastases":
            return ("Is there likely or definintely 1 or more metastases to the liver (do not include other organs)? Options are: ['Yes', 'No'].  If not specifically mentioned, then answer 'No'") 
        if exam_class.lower() == "glioma progression":
            return ("What changes are seen in the brain tumor compared to only the most recent examination? Options are: ['Progression', 'Stable', 'Improved', 'Pseudoprogression', 'pseudoreponse']. Usually, increase in size means progression, and decrease in size means improved, but pseudoprogression and pseudoresponse can be exceptions to this. if there is no clear change in the tumor except for post-operative changes or if tumor status is not mentioned, then it is Stable.") 
        return None
        
    
    #---
    
    def get_answer(text, question):
        openai.api_key = 'lm-studio'
    
        content = f"Use the following radiology report to answer the question that follows. Note the 'IMPRESSION' section usually has the most reliable information: {text} Question: {question}. Answer with respect to the current examination and ignore prior exam findings."
        
        history = [
        {"role": "system", "content": "You are an expert radiologist and research. You always provide precise single word answers to the question. When given a list of choices, you only select answers from teh list of choices. If you are not sure, you select the best of the options. You can think through the options after arriving at an initial answer, but the first word of your response is in the list. "},
    #    {"role": "user", "content": "You will receive a radiology report followed by a question. Please answer that question using only the options listed"},
        {"role": "user", "content": content}
    ]
        completion = client.chat.completions.create(
            model=MODEL,
            messages=history,
            temperature=0.05,
            stream=False,
        )
        
        response = completion.choices[0].message.content   #['choices'][0]['message']['content']
        # print (completion)
        return response
    
    #---
    
    #   if len(sys.argv) != 2:
    #       print("Usage: python answer_questions.py <excel-file>")
    #       return
    
    #    workbook = load_workbook(filename=sys.argv[1])
    filename = "./SIIMReports.xlsx"
    workbook = load_workbook(filename=filename)
    sheet = workbook['Sheet1']  # Assuming the questions are in column A and answers will be written to column B
    
    num_correct = 0
    num_incorrect = 0
    
    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):  # Assuming the first row contains headers
        if i < 1000:  # in case you want to run a short subset to test
            report = row[0]  # report must be in first column
            truth = row[1]  # correct answer must be in 2nd column
            exam_class = row[2] # which of the 5 categories in 3rd column
            report = report.replace('_x000D_', '')
            re.sub(r'\r?\n',' ',report)
            re.sub(r' +',' ',report).strip()
            sheet.cell(row=i+2, column=1).value = '<deleted>'  # delete report 
            if len(exam_class) > -1:
                #$print(f"{i}--Report'{report[:55]}' Answer: {answer} exam class '{exam_class}'")
                question = get_question(exam_class)
                if question != None:
                #answer = answer_question(question)
                    answer = get_answer(report, question)
                    sheet.cell(row=i+2, column=1).value = answer  # Put answers back where report was 
                    if answer.lower() == truth.lower():
                        correct = 1
                        num_correct += 1
                    else:
                        correct = 0
                        num_incorrect += 1
                    if ((i+2) % 10) == 0:
                        print (f"set row {i+2} to {answer} vs truth: {truth} so correct = {correct}")
                    sheet.cell(row=i+2, column=4).value = correct
                else:
                    print (f"No question: {question}  Exam Class: {exam_class}")
    
    sheet.cell(row=1, column=1).value = 'Prediction'
    sheet.cell(row=1, column=4).value = 'Correct=1'
    sheet.cell(row=1, column=5).value = MODEL
    
    outfilename = "./SIIMResults.xlsx"
    
    print (f"Overall got {num_correct} out of {num_incorrect + num_correct} right.")
    #sheet.delete_cols(1,1)
    workbook.save(filename=outfilename)


##########################################################################
# This file was converted using nb2py: https://github.com/BardiaKh/nb2py #
##########################################################################
